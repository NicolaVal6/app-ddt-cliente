import streamlit as st
import pandas as pd
import io
import zipfile
import re
import os
import math
import textwrap
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURAZIONE DI SISTEMA ---
st.set_page_config(page_title="Trasporti App v8.9", layout="centered")
st.caption("Versione Sistema: 8.9 (Doppio Foglio Excel per Dati Extra e Viaggi)")

PASSWORD_CLIENTE = "Trasporti2024!"

def check_password():
    if "password_corretta" not in st.session_state:
        st.session_state["password_corretta"] = False
    if not st.session_state["password_corretta"]:
        st.warning("🔒 Area Riservata. Inserisci la password.")
        pwd_inserita = st.text_input("Password", type="password")
        if st.button("Accedi"):
            if pwd_inserita == PASSWORD_CLIENTE:
                st.session_state["password_corretta"] = True
                st.rerun()
            else:
                st.error("Password errata.")
        return False
    return True

if not check_password():
    st.stop()


# --- MOTORE DI FORMATTAZIONE AVANZATO EXCEL ---
def formatta_excel_valsecchi(writer, sheet_name, is_casati=False, cliente_nome="", info_anagrafica="", tipo_foglio="viaggi"):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    font_bold_standard = Font(name='Calibri', size=11, bold=True, color="000000")
    font_bold_calibri12 = Font(name='Calibri', size=12, bold=True, color="000000")
    
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    fill_azzurrino = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )

    start_row_header = 7 if is_casati else 1

    # Pre-impostazioni di stampa A4 orizzontale in background
    worksheet.page_setup.orientation = 'landscape' 
    worksheet.page_setup.paperSize = 9  
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = False

    # Configurazione Margini di Sicurezza Piè di Pagina
    worksheet.oddFooter.center.text = "Pagina &P"
    worksheet.page_margins.left = 0.3
    worksheet.page_margins.right = 0.3
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.9  
    worksheet.page_margins.footer = 0.4  

    if is_casati:
        # =========================================================================
        # 📑 BI-REGOLAMENTO DI RIPETIZIONE RIGHE (v8.9)
        # =========================================================================
        if tipo_foglio == "viaggi":
            worksheet.print_title_rows = '1:7' # Ripete anche i titoli colonne dei viaggi
        else:
            worksheet.print_title_rows = '1:6' # Nel foglio extra esclude la riga 7 (Nessun testo fantasma)
        # =========================================================================
        
        worksheet.row_dimensions[1].height = 80
        
        logo_path = None
        for filename in ["logo.png", "logo.jpg", "logo.jpeg", "LOGO.jpg", "LOGO.PNG"]:
            if os.path.exists(filename):
                logo_path = filename
                break
        
        if logo_path:
            try:
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img = OpenpyxlImage(logo_path)
                img.width = 147  
                img.height = 91  
                worksheet.add_image(img, 'I1')
            except:
                pass
        
        worksheet.row_dimensions[2].height = 65 
        cell_azienda = worksheet['A2']
        cell_azienda.value = "VALSECCHI TRASPORTI S.R.L.\nVIA GIUSEPPE PARINI N. 8\nCESANA BRIANZA (LC)\nP.IVA 01932020132"
        cell_azienda.font = font_bold_calibri12
        cell_azienda.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        testo_spettabile = info_anagrafica if info_anagrafica else cliente_nome
        if info_anagrafica:
            parti_indirizzo = [p.strip() for p in re.split(r'\s{2,}', str(info_anagrafica)) if p.strip()]
            if parti_indirizzo:
                parti_indirizzo[0] = f"SPETT.LE   {parti_indirizzo[0]}"
                testo_finito = "\n".join(parti_indirizzo)
            else:
                testo_finito = f"SPETT.LE   {info_anagrafica}"
        else:
            testo_finito = f"SPETT.LE   {cliente_nome}"

        cell_spett = worksheet['A4']
        cell_spett.value = testo_finito
        cell_spett.font = font_bold_calibri12
        cell_spett.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.row_dimensions[4].height = 65 
        
        cell_fatt = worksheet['A5']
        cell_fatt.value = "RIF. NS FATT. N. ........ DEL  ..../..../202.."
        cell_fatt.font = font_bold_standard
        cell_fatt.alignment = align_left_wrap
        worksheet.row_dimensions[5].height = 18

    # Gestione geometrica larghezze colonne differenziata per foglio
    if tipo_foglio == "extra":
        worksheet.column_dimensions['A'].width = 30.9
        for col_idx in range(2, 10):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 16
    else:
        for col_idx in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = str(worksheet.cell(row=start_row_header, column=col_idx).value).upper()
            
            max_len = 0
            for row_idx in range(start_row_header, worksheet.max_row + 1):
                val = worksheet.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            
            max_allowed_width = 25 
            if col_letter == 'A' and is_casati:
                worksheet.column_dimensions[col_letter].width = 30.9
            else:
                if max_len > max_allowed_width and any(x in header_val for x in ['DITTA', 'LUOGO', 'DESTINAZIONE', 'VARIE']):
                    worksheet.column_dimensions[col_letter].width = max_allowed_width
                else:
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # Formattazione allineamenti base della griglia (attiva solo se ci sono dati)
    for row in worksheet.iter_rows(min_row=start_row_header):
        for cell in row:
            cell.font = font_bold_standard
            cell.border = thin_border
            if cell.column <= 4:
                cell.alignment = align_left_wrap
            else:
                cell.alignment = align_right_wrap

    # Motore di simulazione altezze righe (TextWrap)
    start_dynamic_row = 4 if is_casati else 1
    for row_idx in range(start_dynamic_row, worksheet.max_row + 1):
        max_lines_in_row = 0
        has_text = False
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                has_text = True
                val_str = str(cell.value)
                col_letter = get_column_letter(col_idx)
                
                w = worksheet.column_dimensions[col_letter].width or 16
                safe_width = max(5, int(w) - 2)
                
                lines_for_this_cell = 0
                for paragraph in val_str.split('\n'):
                    wrapped = textwrap.wrap(paragraph, width=safe_width)
                    lines_for_this_cell += len(wrapped) if wrapped else 1
                
                if lines_for_this_cell > max_lines_in_row:
                    max_lines_in_row = lines_for_this_cell
        
        if has_text:
            if row_idx == 4 and is_casati:
                worksheet.row_dimensions[row_idx].height = max(65, max_lines_in_row * 16.5)
            elif row_idx < start_row_header:
                worksheet.row_dimensions[row_idx].height = max(18, max_lines_in_row * 16.5)
            else:
                worksheet.row_dimensions[row_idx].height = max(24.9, max_lines_in_row * 16.5)

    # Blocchi di formattazione specifici per i fogli contenenti tabelle reali
    if tipo_foglio != "extra":
        # Sfondo azzurrino riga intestazione
        for cell in worksheet[start_row_header]:
            cell.fill = fill_azzurrino

        # Formattazione Decimali e Date
        for col_idx in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = str(worksheet.cell(row=start_row_header, column=col_idx).value).upper()
            
            for row_idx in range(start_row_header + 1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    if any(x in header_val for x in ['TARIFFA', 'TOTALE']):
                        cell.number_format = '#,##0.00'
                    if any(x in header_val for x in ['DATA', 'DDT']) and not isinstance(cell.value, str):
                        cell.number_format = 'DD/MM/YYYY'

        # Calcolo Totale Automatico
        totale_col_idx = None
        for col_idx in range(1, worksheet.max_column + 1):
            if "TOTALE" in str(worksheet.cell(row=start_row_header, column=col_idx).value).upper():
                totale_col_idx = col_idx
                break
        
        if totale_col_idx:
            last_data_row = worksheet.max_row
            total_row_idx = last_data_row + 1
            
            worksheet.row_dimensions[total_row_idx].height = 24.9
            cell_label = worksheet.cell(row=total_row_idx, column=1)
            cell_label.value = "TOTALE"
            
            col_letter = get_column_letter(totale_col_idx)
            start_data_row = 8 if is_casati else 2 
            
            cell_sum = worksheet.cell(row=total_row_idx, column=totale_col_idx)
            cell_sum.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{last_data_row})"
            cell_sum.number_format = '#,##0.00'
            
            for col_idx in range(1, worksheet.max_column + 1):
                c = worksheet.cell(row=total_row_idx, column=col_idx)
                c.font = font_bold_standard
                c.border = thin_border
                if col_idx <= 4:
                    c.alignment = align_left_wrap
                else:
                    c.alignment = align_right_wrap


# --- INTERFACCIA WEB ---
try: st.image("logo.png" if os.path.exists("logo.png") else "LOGO.jpg", width=250)
except: pass

st.title("🚛 Sistema Gestionale Valsecchi Trasporti")

uploaded_file = st.file_uploader("1. Carica Estrazione Gestionale (.xlsx o .csv)", type=["xlsx", "csv"])
uploaded_anagrafica = st.file_uploader("2. Carica Anagrafiche Clienti (Opzionale)", type=["xlsx", "csv"])

if uploaded_file and st.button("🚀 GENERA DOCUMENTI FISCALI", type="primary"):
    try:
        anagrafiche = {}
        if uploaded_anagrafica:
            df_ana = pd.read_excel(uploaded_anagrafica) if uploaded_anagrafica.name.endswith('.xlsx') else pd.read_csv(uploaded_anagrafica)
            df_ana.columns = df_ana.columns.str.strip().str.upper()
            
            col_target = None
            for c in df_ana.columns:
                if 'INTESTAZIONE' in c:
                    col_target = c
                    break
            
            if 'CLIENTE' in df_ana.columns and col_target:
                anagrafiche = dict(zip(
                    df_ana['CLIENTE'].astype(str).str.upper().str.strip(), 
                    df_ana[col_target].astype(str).str.strip()
                ))

        if uploaded_file.name.endswith('.xlsx'):
            hdr = pd.read_excel(uploaded_file, nrows=0).columns
            uploaded_file.seek(0)
            dtypes = {c: str for c in hdr if 'DDT' in c.upper() and 'DATA' not in c.upper()}
            df = pd.read_excel(uploaded_file, dtype=dtypes)
        else:
            hdr = pd.read_csv(uploaded_file, nrows=0).columns
            uploaded_file.seek(0)
            dtypes = {c: str for c in hdr if 'DDT' in c.upper() and 'DATA' not in c.upper()}
            df = pd.read_csv(uploaded_file, dtype=dtypes)
            
        df.columns = df.columns.str.strip().str.upper()
        df = df.loc[:, ~df.columns.str.contains('^UNNAMED', na=False)]

        mappa = {
            'DITTA PRESA': 'DITTA PRESA', 'DITTA DI CARICO': 'DITTA PRESA',
            'LUOGO PRESA': 'LUOGO PRESA', 'LUOGO DI CARICO': 'LUOGO PRESA',
            'DESTINAZIONE FINALE': 'DESTINAZIONE FINALE', 'DESTINAZIONE': 'DESTINAZIONE FINALE',
            'TIPO MERCE': 'VARIE', 'VARIE': 'VARIE',
            'QUANTITA': "QUANTITA'", 'PESO': "QUANTITA'",
            'DDT': 'N. DDT', 'N. DDT': 'N. DDT',
            'TARIFFA': 'TARIFFA', 'TOTALE': 'TOTALE', 'DATA DDT': 'DATA DDT'
        }
        df_mapped = df.rename(columns=mappa)
        
        if 'DATA DDT' in df_mapped.columns:
            df_mapped['DATA DDT'] = pd.to_datetime(df_mapped['DATA DDT'], errors='coerce').dt.date

        cols_ordinate = ['DITTA PRESA', 'LUOGO PRESA', 'DESTINAZIONE FINALE', 'VARIE', "QUANTITA'", 'TARIFFA', 'TOTALE', 'N. DDT', 'DATA DDT']
        cols_presenti = [c for c in cols_ordinate if c in df_mapped.columns]

        first_valid = df_mapped['DATA DDT'].dropna() if 'DATA DDT' in df_mapped.columns else pd.Series()
        mese_str = first_valid.iloc[0].strftime("%Y-%m") if not first_valid.empty else "MESE_IGNOTO"

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            
            # --- 1. FILE CLIENTI ---
            for cliente, group in df_mapped.groupby('CLIENTE'):
                safe_cliente = re.sub(r'[\\/*?:"<>|]', "", str(cliente)).strip()
                buf = io.BytesIO()
                
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    data_export = group[cols_presenti]
                    # FOGLIO 1: Report Viaggi standard
                    data_export.to_excel(writer, index=False, sheet_name='Prospetto_DDT', startrow=6)
                    info = anagrafiche.get(str(cliente).upper().strip(), "")
                    formatta_excel_valsecchi(writer, 'Prospetto_DDT', is_casati=True, cliente_nome=str(cliente), info_anagrafica=info, tipo_foglio="viaggi")
                    
                    # FOGLIO 2: Generazione automatica Tab Extra Vuoto (con solo intestazione)
                    workbook = writer.book
                    worksheet_extra = workbook.create_sheet(title='Dati_Extra')
                    writer.sheets['Dati_Extra'] = worksheet_extra
                    formatta_excel_valsecchi(writer, 'Dati_Extra', is_casati=True, cliente_nome=str(cliente), info_anagrafica=info, tipo_foglio="extra")
                
                zip_file.writestr(f"Clienti/{safe_cliente}_{mese_str}.xlsx", buf.getvalue())

            # --- 2. FILE AUTISTI ---
            if 'AUTISTA ALLO SCARICO' in df.columns:
                for autista, group in df.groupby('AUTISTA ALLO SCARICO'):
                    safe_autista = re.sub(r'[\\/*?:"<>|]', "", str(autista)).strip()
                    buf = io.BytesIO()
                    
                    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                        group.to_excel(writer, index=False, sheet_name='Prospetto_Autista')
                        formatta_excel_valsecchi(writer, 'Prospetto_Autista', is_casati=False, tipo_foglio="viaggi")
                        
                    zip_file.writestr(f"Autisti/Scarico_{safe_autista}_{mese_str}.xlsx", buf.getvalue())

        st.success("✅ File bi-foglio strutturati con successo! Pronto per l'esportazione PDF unificata.")
        st.download_button("📥 SCARICA ARCHIVIO COMPLETO", zip_buffer.getvalue(), f"Trasporti_Valsecchi_{mese_str}.zip", "application/zip")

    except Exception as e:
        st.error(f"Errore: {e}")
