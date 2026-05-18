import streamlit as st
import pandas as pd
import io
import zipfile
import re
import os
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURAZIONE DI SISTEMA ---
st.set_page_config(page_title="Trasporti App v6.0", layout="centered")
st.caption("Versione Sistema: 6.0 (Specifiche Valsecchi Trasporti)")

PASSWORD_CLIENTE = "Trasporti2024!"

def check_password():
    if "password_corretta" not in st.session_state:
        st.session_state["password_corretta"] = False
    if not st.session_state["password_corretta"]:
        st.warning("🔒 Area Riservata. Inserisci la password.")
        pwd_inserita = st.text_input("Password", type="password")
        if st.button("Accedi"):
            if pwd_inserita == PASSWORD_CLIENTE:
                st.session_state["password_corretta"] = True
                st.rerun()
            else:
                st.error("Password errata.")
        return False
    return True

if not check_password():
    st.stop()


# --- MOTORE DI FORMATTAZIONE AVANZATO EXCEL ---
def formatta_excel_valsecchi(writer, sheet_name, is_casati=False, cliente_nome="", info_anagrafica=""):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Stili Rigorosi (Tutto in Grassetto)
    font_bold_standard = Font(name='Calibri', size=11, bold=True, color="000000")
    font_bold_grande = Font(name='Calibri', size=12, bold=True, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
    )

    start_row_header = 6 if is_casati else 1
    
    # Configurazione Piè di pagina (Numero di pagina)
    worksheet.page_footer.center.text = "Pagina &P"

    if is_casati:
        # Ripetizione Intestazioni su ogni pagina stampata
        worksheet.print_title_rows = '1:6'
        worksheet.row_dimensions[2].height = 22
        
        # 1. Intestazione Aziendale AUTOMATICA A DESTRA
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
        cell_azienda = worksheet.cell(row=2, column=1)
        cell_azienda.value = "VALSECCHI TRASPORTI S.R.L.   VIA GIUSEPPE PARINI N. 8   CESANA BRIANZA (LC)   P.IVA 01932020132"
        cell_azienda.font = font_bold_grande
        cell_azienda.alignment = align_right
        
        # Iniezione Logo a Sinistra (se presente nel server)
        if os.path.exists("logo.png"):
            try:
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img = OpenpyxlImage("logo.png")
                img.width = 110
                img.height = 40
                worksheet.add_image(img, 'A1')
            except:
                pass
        
        # 2. Scritta Spettabile a Sinistra
        testo_spettabile = info_anagrafica if info_anagrafica else cliente_nome
        worksheet['A4'] = f"SPETT.LE   {testo_spettabile}"
        worksheet['A4'].font = font_bold_standard
        worksheet['A4'].alignment = align_left
        
        # 3. Riferimento Fattura Fissa
        worksheet['A5'] = "RIF. NS FATT. N. ........ DEL  ..../..../202.."
        worksheet['A5'].font = font_bold_standard
        worksheet['A5'].alignment = align_left

    # 4. Formattazione Griglia Dati (Tutto Grassetto, Tutto Centrato)
    for row in worksheet.iter_rows(min_row=start_row_header):
        for cell in row:
            cell.font = font_bold_standard
            cell.border = thin_border
            cell.alignment = align_center

    # 5. Formattazione Numerica (Due decimali) e Date
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_val = str(worksheet.cell(row=start_row_header, column=col_idx).value).upper()
        
        for row_idx in range(start_row_header + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                if any(x in header_val for x in ['TARIFFA', 'TOTALE']):
                    cell.number_format = '#,##0.00'
                if any(x in header_val for x in ['DATA', 'DDT']) and not isinstance(cell.value, str):
                    cell.number_format = 'DD/MM/YYYY'
        
        # Auto-fit colonne proporzionato
        max_len = max(len(str(cell.value or '')) for cell in worksheet[col_letter] if cell.row >= start_row_header)
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # 6. CALCOLO AUTOMATICO DEL TOTALE DELLE SOMME (Due decimali)
    if is_casati:
        totale_col_idx = None
        for col_idx in range(1, worksheet.max_column + 1):
            if "TOTALE" in str(worksheet.cell(row=start_row_header, column=col_idx).value).upper():
                totale_col_idx = col_idx
                break
        
        if totale_col_idx:
            last_data_row = worksheet.max_row
            total_row_idx = last_data_row + 1
            
            # Scritta TOTALE sulla prima colonna
            cell_label = worksheet.cell(row=total_row_idx, column=1)
            cell_label.value = "TOTALE"
            
            # Formula di somma automatica Excel
            col_letter = get_column_letter(totale_col_idx)
            cell_sum = worksheet.cell(row=total_row_idx, column=totale_col_idx)
            cell_sum.value = f"=SUM({col_letter}7:{col_letter}{last_data_row})"
            cell_sum.number_format = '#,##0.00'
            
            # Applica stile uniforme alla riga dei totali
            for col_idx in range(1, worksheet.max_column + 1):
                c = worksheet.cell(row=total_row_idx, column=col_idx)
                c.font = font_bold_standard
                c.border = thin_border
                c.alignment = align_center


# --- INTERFACCIA WEB ---
try: st.image("logo.png", width=250)
except: pass

st.title("🚛 Sistema Gestionale Valsecchi Trasporti")

uploaded_file = st.file_uploader("1. Carica Estrazione Gestionale (.xlsx o .csv)", type=["xlsx", "csv"])
uploaded_anagrafica = st.file_uploader("2. Carica Anagrafiche Clienti (Opzionale)", type=["xlsx", "csv"])

if uploaded_file and st.button("🚀 GENERA DOCUMENTI FISCALI", type="primary"):
    try:
        # Parsing Anagrafiche
        anagrafiche = {}
        if uploaded_anagrafica:
            df_ana = pd.read_excel(uploaded_anagrafica) if uploaded_anagrafica.name.endswith('.xlsx') else pd.read_csv(uploaded_anagrafica)
            df_ana.columns = df_ana.columns.str.strip().str.upper()
            if 'CLIENTE' in df_ana.columns and 'INTESTAZIONE COMPLETA' in df_ana.columns:
                anagrafiche = dict(zip(df_ana['CLIENTE'].astype(str).str.upper(), df_ana['INTESTAZIONE COMPLETA']))

        # Lettura file trasporti
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith('.xlsx') else pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip().str.upper()
        df = df.loc[:, ~df.columns.str.contains('^UNNAMED', na=False)]

        # Traduzione dizionario campi
        mappa = {
            'DITTA PRESA': 'DITTA PRESA', 'DITTA DI CARICO': 'DITTA PRESA',
            'LUOGO PRESA': 'LUOGO PRESA', 'LUOGO DI CARICO': 'LUOGO PRESA',
            'DESTINAZIONE FINALE': 'DESTINAZIONE FINALE', 'DESTINAZIONE': 'DESTINAZIONE FINALE',
            'TIPO MERCE': 'VARIE', 'VARIE': 'VARIE',
            'QUANTITA': "QUANTITA'", 'PESO': "QUANTITA'",
            'DDT': 'N. DDT', 'N. DDT': 'N. DDT',
            'TARIFFA': 'TARIFFA', 'TOTALE': 'TOTALE', 'DATA DDT': 'DATA DDT'
        }
        df_mapped = df.rename(columns=mappa)
        
        if 'DATA DDT' in df_mapped.columns:
            df_mapped['DATA DDT'] = pd.to_datetime(df_mapped['DATA DDT'], errors='coerce').dt.date

        # CONFIGURAZIONE RIGOROSA DELL'ORDINE COLONNE RICHIESTO (Sinistra -> Destra)
        cols_ordinate = ["QUANTITA'", 'TARIFFA', 'TOTALE', 'N. DDT', 'DATA DDT', 'DITTA PRESA', 'LUOGO PRESA', 'DESTINAZIONE FINALE', 'VARIE']
        cols_presenti = [c for c in cols_ordinate if c in df_mapped.columns]

        first_valid = df_mapped['DATA DDT'].dropna() if 'DATA DDT' in df_mapped.columns else pd.Series()
        mese_str = first_valid.iloc[0].strftime("%Y-%m") if not first_valid.empty else "MESE_IGNOTO"

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            
            # --- 1. GENERAZIONE EXCEL CLIENTI ---
            for cliente, group in df_mapped.groupby('CLIENTE'):
                safe_cliente = re.sub(r'[\\/*?:"<>|]', "", str(cliente)).strip()
                buf = io.BytesIO()
                
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    data_export = group[cols_presenti]
                    # startrow=5 inserisce le intestazioni alla riga 6 di Excel
                    data_export.to_excel(writer, index=False, sheet_name='Prospetto_DDT', startrow=5)
                    
                    info = anagrafiche.get(str(cliente).upper(), "")
                    formatta_excel_valsecchi(writer, 'Prospetto_DDT', is_casati=True, cliente_nome=str(cliente), info_anagrafica=info)
                
                zip_file.writestr(f"Clienti/{safe_cliente}_{mese_str}.xlsx", buf.getvalue())

            # --- 2. GENERAZIONE EXCEL AUTISTI CON FILTRO ALLO SCARICO ---
            if 'AUTISTA ALLO SCARICO' in df.columns:
                for autista, group in df.groupby('AUTISTA ALLO SCARICO'):
                    safe_autista = re.sub(r'[\\/*?:"<>|]', "", str(autista)).strip()
                    buf = io.BytesIO()
                    
                    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                        group.to_excel(writer, index=False, sheet_name='Prospetto_Autista')
                        formatta_excel_valsecchi(writer, 'Prospetto_Autista', is_casati=False)
                    
                    zip_file.writestr(f"Autisti/Scarico_{safe_autista}_{mese_str}.xlsx", buf.getvalue())
            else:
                st.warning("⚠️ Colonna 'AUTISTA ALLO SCARICO' non trovata nel file. Sezione autisti saltata.")

        st.success("✅ Elaborazione completata secondo le specifiche Valsecchi Trasporti!")
        st.download_button("📥 SCARICA ARCHIVIO COMPLETO", zip_buffer.getvalue(), f"Trasporti_Valsecchi_{mese_str}.zip", "application/zip")

    except Exception as e:
        st.error(f"Errore durante l'esecuzione del codice: {e}")
